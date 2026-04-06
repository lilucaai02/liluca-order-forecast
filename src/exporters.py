"""エクスポーター: CSV / Excel / PDF / Google Sheets への出力."""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Optional

from src.models import InventoryItem, ThresholdConfig

if TYPE_CHECKING:
    from src.forecasting import ForecastResult


# ---------------------------------------------------------------------------
#  共通ヘルパー
# ---------------------------------------------------------------------------

def _status_label(qty: int, threshold: ThresholdConfig) -> str:
    """在庫状態のラベルを返す."""
    if qty == 0:
        return "在庫切れ"
    if qty <= threshold.critical_level:
        return "危険"
    if qty <= threshold.reorder_point:
        return "要発注"
    return "正常"


def _inventory_rows(
    items: list[InventoryItem],
    defaults: ThresholdConfig,
    sku_configs: dict[str, ThresholdConfig],
) -> list[dict]:
    """InventoryItem リストを辞書のリストに変換."""
    rows = []
    for item in items:
        threshold = sku_configs.get(item.seller_sku, defaults)
        inbound = item.inbound_shipped_quantity + item.inbound_receiving_quantity
        mp_label = "Amazon" if item.marketplace == "amazon" else "楽天"
        rows.append({
            "MP": mp_label,
            "アカウント": item.account_name,
            "SKU": item.seller_sku,
            "ASIN": item.asin,
            "商品名": item.product_name,
            "出荷可能": item.fulfillable_quantity,
            "予約済": item.reserved_quantity,
            "入庫中": inbound,
            "合計": item.total_quantity,
            "発注点": threshold.reorder_point,
            "状態": _status_label(item.fulfillable_quantity, threshold),
        })
    return rows


def _forecast_rows(results: list[ForecastResult]) -> list[dict]:
    """ForecastResult リストを辞書のリストに変換."""
    rows = []
    for r in results:
        rows.append({
            "SKU": r.seller_sku,
            "ASIN": r.asin,
            "商品名": r.product_name,
            "現在庫": r.current_stock,
            "日次消費": round(r.daily_consumption_rate, 1),
            "在庫切れまで(日)": r.days_until_stockout if r.days_until_stockout is not None else "",
            "発注点到達(日)": r.days_until_reorder_point if r.days_until_reorder_point is not None else "",
            "推奨発注数": r.recommended_order_qty if r.recommended_order_qty > 0 else "不要",
            "発注期限": r.recommended_order_date or "",
            "分析方法": r.method,
        })
    return rows


# ---------------------------------------------------------------------------
#  CSV エクスポート
# ---------------------------------------------------------------------------

def export_csv(
    items: list[InventoryItem],
    defaults: ThresholdConfig,
    sku_configs: dict[str, ThresholdConfig],
    filepath: Path,
    *,
    forecast_results: Optional[list[ForecastResult]] = None,
) -> Path:
    """在庫データをCSVファイルにエクスポート.

    forecast_results が指定されている場合は予測データのCSVを出力する。
    """
    if forecast_results:
        rows = _forecast_rows(forecast_results)
    else:
        rows = _inventory_rows(items, defaults, sku_configs)

    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    return filepath


# ---------------------------------------------------------------------------
#  Excel エクスポート
# ---------------------------------------------------------------------------

def export_excel(
    items: list[InventoryItem],
    defaults: ThresholdConfig,
    sku_configs: dict[str, ThresholdConfig],
    filepath: Path,
    *,
    forecast_results: Optional[list[ForecastResult]] = None,
) -> Path:
    """在庫データをExcelファイルにエクスポート (色付き)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    # === 在庫シート ===
    ws_inv = wb.active
    ws_inv.title = "在庫状況"
    inv_rows = _inventory_rows(items, defaults, sku_configs)
    _write_excel_sheet(ws_inv, inv_rows, "FBA在庫状況", _status_fill_map())

    # === 予測シート (あれば) ===
    if forecast_results:
        ws_fc = wb.create_sheet("発注予測")
        fc_rows = _forecast_rows(forecast_results)
        _write_excel_sheet(ws_fc, fc_rows, "発注予測")

    wb.save(str(filepath))
    return filepath


def _status_fill_map() -> dict:
    """状態セルの色マッピング."""
    from openpyxl.styles import PatternFill

    return {
        "在庫切れ": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "危険": PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid"),
        "要発注": PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
        "正常": PatternFill(start_color="44BB44", end_color="44BB44", fill_type="solid"),
    }


def _write_excel_sheet(
    ws,
    rows: list[dict],
    title: str,
    status_fills: Optional[dict] = None,
) -> None:
    """ワークシートにデータを書き込み、スタイリングを適用."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if not rows:
        return

    headers = list(rows[0].keys())
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    # タイトル行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"{title}  ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # ヘッダー行
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # データ行
    for row_idx, row_data in enumerate(rows, start=4):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # 数値列は右寄せ
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

            # 状態列に色付け
            if header == "状態" and status_fills and value in status_fills:
                cell.fill = status_fills[value]
                cell.font = Font(color="FFFFFF", bold=True)

    # 列幅の自動調整
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_data in rows:
            val_len = len(str(row_data[header]))
            if val_len > max_len:
                max_len = val_len
        # 日本語文字幅を考慮 (おおよそ2倍)
        adjusted = min(max_len * 2 + 4, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # フィルター
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(rows)}"
    # ウィンドウ枠の固定
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
#  PDF レポート
# ---------------------------------------------------------------------------

_FONT_PATHS = [
    "/Library/Fonts/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/usr/share/fonts/truetype/noto/NotoSansCJKjp-Regular.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
]

_FONT_FAMILY = "ArialUni"


def _find_unicode_font() -> str | None:
    """利用可能な日本語対応TTFフォントパスを返す."""
    import os
    for p in _FONT_PATHS:
        if os.path.isfile(p):
            return p
    return None


def export_pdf(
    items: list[InventoryItem],
    defaults: ThresholdConfig,
    sku_configs: dict[str, ThresholdConfig],
    filepath: Path,
    *,
    forecast_results: Optional[list[ForecastResult]] = None,
    sales_summary: Optional[dict] = None,
) -> Path:
    """在庫レポートをPDFファイルにエクスポート."""
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    pdf = _make_pdf()
    pdf.add_page()

    font = _FONT_FAMILY if _find_unicode_font() else "Helvetica"

    # --- サマリーセクション ---
    total_skus = len(items)
    out_of_stock = sum(1 for i in items if i.fulfillable_quantity == 0)
    critical = sum(
        1 for i in items
        if 0 < i.fulfillable_quantity <= sku_configs.get(
            i.seller_sku, defaults
        ).critical_level
    )
    warning = sum(
        1 for i in items
        if sku_configs.get(i.seller_sku, defaults).critical_level
        < i.fulfillable_quantity
        <= sku_configs.get(i.seller_sku, defaults).reorder_point
    )
    healthy = total_skus - out_of_stock - critical - warning

    pdf.section_title("Summary")
    pdf.set_font(font, "", 10)

    summary_data = [
        ["SKU Total", str(total_skus)],
        ["Out of Stock", str(out_of_stock)],
        ["Critical", str(critical)],
        ["Warning", str(warning)],
        ["Healthy", str(healthy)],
    ]
    if sales_summary:
        summary_data.append(["Daily Sales (avg)", f"{sales_summary.get('avg_daily', 0):.1f} units"])
        summary_data.append(["Period Sales", f"{sales_summary.get('total_units', 0)} units"])

    col_w = [50, 40]
    for row in summary_data:
        pdf.cell(col_w[0], 7, row[0], border=0)
        pdf.cell(col_w[1], 7, row[1], border=0, new_x="LMARGIN", new_y="NEXT")

    pdf.ln(5)

    # --- 在庫テーブル ---
    pdf.section_title("Inventory Status")
    inv_rows = _inventory_rows(items, defaults, sku_configs)
    inv_headers = ["SKU", "ASIN", "Product", "Available", "Total", "Reorder Pt", "Status"]
    inv_col_w = [30, 25, 52, 18, 15, 20, 18]

    # ヘッダー
    pdf.set_fill_color(43, 87, 151)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(font, "B", 7)
    for i, h in enumerate(inv_headers):
        pdf.cell(inv_col_w[i], 6, h, border=1, fill=True, align="C")
    pdf.ln()

    # データ
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(font, "", 7)
    for row in inv_rows:
        # 状態に応じた行色
        status = row["状態"]
        if status == "在庫切れ":
            pdf.set_fill_color(255, 200, 200)
        elif status == "危険":
            pdf.set_fill_color(255, 220, 180)
        elif status == "要発注":
            pdf.set_fill_color(255, 240, 180)
        else:
            pdf.set_fill_color(220, 255, 220)

        values = [
            str(row["SKU"])[:18],
            str(row["ASIN"]),
            str(row["商品名"])[:28],
            str(row["出荷可能"]),
            str(row["合計"]),
            str(row["発注点"]),
            status,
        ]

        for i, val in enumerate(values):
            align = "R" if i in (3, 4, 5) else "L"
            pdf.cell(inv_col_w[i], 5.5, val, border=1, fill=True, align=align)
        pdf.ln()

        # ページ下端チェック
        if pdf.get_y() > 180:
            pdf.add_page()

    # --- 発注予測テーブル (あれば) ---
    if forecast_results:
        pdf.ln(5)
        pdf.section_title("Forecast & Reorder Recommendations")
        fc_rows = _forecast_rows(forecast_results)
        fc_headers = ["SKU", "Product", "Stock", "Daily Use", "Stockout", "Reorder Qty", "Deadline"]
        fc_col_w = [30, 50, 16, 18, 18, 22, 24]

        pdf.set_fill_color(43, 87, 151)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font(font, "B", 7)
        for i, h in enumerate(fc_headers):
            pdf.cell(fc_col_w[i], 6, h, border=1, fill=True, align="C")
        pdf.ln()

        pdf.set_text_color(0, 0, 0)
        pdf.set_font(font, "", 7)
        for row in fc_rows:
            pdf.set_fill_color(245, 245, 255)
            values = [
                str(row["SKU"])[:18],
                str(row["商品名"])[:28],
                str(row["現在庫"]),
                str(row["日次消費"]),
                str(row["在庫切れまで(日)"]),
                str(row["推奨発注数"]),
                str(row["発注期限"]),
            ]
            for i, val in enumerate(values):
                align = "R" if i in (2, 3, 4, 5) else "L"
                pdf.cell(fc_col_w[i], 5.5, val, border=1, fill=True, align=align)
            pdf.ln()

            if pdf.get_y() > 180:
                pdf.add_page()

    pdf.output(str(filepath))
    return filepath


def _make_pdf():
    """日本語対応PDFオブジェクトを生成."""
    from fpdf import FPDF

    font_path = _find_unicode_font()
    font_name = _FONT_FAMILY

    class _PDF(FPDF):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            if font_path:
                self.add_font(font_name, "", font_path, uni=True)
                self.add_font(font_name, "B", font_path, uni=True)
                self.add_font(font_name, "I", font_path, uni=True)

        def header(self):
            f = font_name if font_path else "Helvetica"
            self.set_font(f, "B", 16)
            self.cell(0, 10, "Amazon Inventory Report", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font(f, "", 9)
            self.cell(0, 5, datetime.now().strftime("%Y-%m-%d %H:%M"), align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(5)

        def footer(self):
            f = font_name if font_path else "Helvetica"
            self.set_y(-15)
            self.set_font(f, "I", 8)
            self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

        def section_title(self, title: str):
            f = font_name if font_path else "Helvetica"
            self.set_font(f, "B", 12)
            self.set_fill_color(230, 230, 240)
            self.cell(0, 8, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
            self.ln(3)

    return _PDF(orientation="L", unit="mm", format="A4")


# ---------------------------------------------------------------------------
#  Google Sheets エクスポート
# ---------------------------------------------------------------------------

def export_google_sheets(
    items: list[InventoryItem],
    defaults: ThresholdConfig,
    sku_configs: dict[str, ThresholdConfig],
    spreadsheet_id: str,
    credentials_file: str,
    *,
    forecast_results: Optional[list[ForecastResult]] = None,
) -> str:
    """在庫データをGoogleスプレッドシートに書き込み.

    Args:
        spreadsheet_id: GoogleスプレッドシートのID (URLの/d/{ID}/部分)
        credentials_file: サービスアカウントJSONキーファイルパス

    Returns:
        スプレッドシートのURL
    """
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(credentials_file, scopes=scopes)
    gc = gspread.authorize(creds)

    spreadsheet = gc.open_by_key(spreadsheet_id)

    # === 在庫シート ===
    inv_rows = _inventory_rows(items, defaults, sku_configs)
    _update_gsheet_worksheet(spreadsheet, "在庫状況", inv_rows)

    # === 予測シート ===
    if forecast_results:
        fc_rows = _forecast_rows(forecast_results)
        _update_gsheet_worksheet(spreadsheet, "発注予測", fc_rows)

    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"


def _update_gsheet_worksheet(spreadsheet, sheet_name: str, rows: list[dict]) -> None:
    """Googleスプレッドシートのワークシートを更新."""
    import gspread

    if not rows:
        return

    # ワークシートの取得または作成
    try:
        ws = spreadsheet.worksheet(sheet_name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=sheet_name, rows=len(rows) + 5, cols=len(rows[0]) + 2)

    headers = list(rows[0].keys())
    # タイトル行
    ws.update_cell(1, 1, f"{sheet_name} (更新: {datetime.now().strftime('%Y-%m-%d %H:%M')})")

    # ヘッダー
    ws.update(range_name="A3", values=[headers])

    # データ
    data = [[row[h] for h in headers] for row in rows]
    ws.update(range_name=f"A4", values=data)

    # ヘッダー行をボールド化 & 固定
    ws.format("A3:Z3", {
        "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        "backgroundColor": {"red": 0.17, "green": 0.34, "blue": 0.59},
        "horizontalAlignment": "CENTER",
    })

    # 状態列に条件付き書式
    if "状態" in headers:
        status_col_idx = headers.index("状態") + 1  # 1-based
        from gspread_formatting import (
            BooleanCondition,
            CellFormat,
            Color,
            ConditionalFormatRule,
            GridRange,
            set_conditional_format_rules,
        )

        rules = []
        color_map = {
            "在庫切れ": Color(1, 0.27, 0.27),
            "危険": Color(1, 0.53, 0),
            "要発注": Color(1, 0.8, 0),
            "正常": Color(0.27, 0.73, 0.27),
        }

        grid = GridRange.from_a1_range(
            f"{chr(64 + status_col_idx)}4:{chr(64 + status_col_idx)}1000",
            ws,
        )
        for text, color in color_map.items():
            rules.append(ConditionalFormatRule(
                ranges=[grid],
                booleanRule=BooleanCondition("TEXT_EQ", [text]),
                format=CellFormat(
                    backgroundColor=color,
                    textFormat={"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                ),
            ))

        try:
            set_conditional_format_rules(ws, rules)
        except Exception:
            pass  # gspread_formatting が無ければスキップ

    # ウィンドウ枠固定
    ws.freeze(rows=3)
